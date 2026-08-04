"""Demo golden lock — monday-morning-report.

The demo deliverable is the committed openpyxl workbook
``output/monday-morning-report.xlsx``. This byte-locks it and asserts the
properties the 07-31 audit cared about:

- The Playfair-Display heading fonts actually stick (the audit flagged a
  regression where the "bold" font had silently reverted). Asserted on BOTH the
  committed workbook AND a freshly generated one — the stated acceptance was
  "assert generated-workbook fonts in a test".
- The three revenue tiers are present (the demo's tier chooser).

If the SHA moves, STOP: the deliverable changed.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "output" / "monday-morning-report.xlsx"

GOLDEN_SHA256 = "555f236c1855e19b05ea9819945d9662248d7094453436c595f23bff9a1d6652"

# Heading cells that must render in Playfair Display (the brand serif).
PLAYFAIR_HEADINGS = [("Setup", "A1"), ("This Week", "A1"), ("Where to Find These", "A1")]


def test_workbook_sha256():
    digest = hashlib.sha256(XLSX.read_bytes()).hexdigest()
    assert digest == GOLDEN_SHA256, (
        f"workbook changed (sha256 {digest} != golden {GOLDEN_SHA256}). "
        "A demo golden moved — STOP and report before re-baselining."
    )


def test_committed_workbook_playfair_headings():
    wb = openpyxl.load_workbook(XLSX)
    for sheet, cell in PLAYFAIR_HEADINGS:
        assert wb[sheet][cell].font.name == "Playfair Display", (
            f"{sheet}!{cell} font is {wb[sheet][cell].font.name!r}, not Playfair Display"
        )


def test_committed_workbook_has_three_tiers():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["_Config"]
    tiers = [ws.cell(row=r, column=1).value for r in range(2, 5)]
    assert all(tiers), "expected 3 tier rows in _Config"
    assert len({t for t in tiers}) == 3


def test_fresh_generate_keeps_playfair(tmp_path, monkeypatch):
    """The Playfair fix must live in generate.py, not just the committed file:
    regenerate and assert the heading fonts are Playfair Display."""
    import generate
    out = tmp_path / "regen.xlsx"
    monkeypatch.setattr(generate, "OUTPUT_PATH", out)
    generate.main()
    wb = openpyxl.load_workbook(out)
    for sheet, cell in PLAYFAIR_HEADINGS:
        assert wb[sheet][cell].font.name == "Playfair Display"
    # tier count preserved
    assert [wb["_Config"].cell(row=r, column=1).value for r in range(2, 5)] == \
        list(__import__("data.metrics", fromlist=["TIER_ORDER"]).TIER_ORDER)
