"""Client-mode CLI for monday-morning-report — the retainer-deliverable generator.

The demo template (generate.py) ships all three revenue tiers behind a chooser —
a site affordance. A retainer deliverable shouldn't ask the client to pick an
audience: client mode ships ONE tier, chosen in engagement.yml, branded for the
client, and (optionally) populated with the client's own KPI values.

The demo path (generate.py, output/monday-morning-report.xlsx) is untouched: this
builds a separate client workbook reusing the tier/metric data in data/metrics.py.

  * tier not one of the three known tiers, or a KPI file whose metrics don't match
    the tier -> a branded Data Readiness Report (HTML), no workbook.
  * otherwise -> a branded, provenance-stamped, DRAFT-until-final client workbook
    (client-output/<engagement>-monday-morning-report.xlsx) with the tier's three
    metrics, their guidance, and the client's values where provided.

Usage:
    python client_mode.py --config engagement.yml [--input client-data/kpis.csv] \
        --out client-output [--final]
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from lailara_engagement import (
    Finding,
    PreflightReport,
    build_provenance,
    load_config,
    write_report,
)
from lailara_engagement.provenance import InputRef

from data.metrics import DEFAULT_NOTE, TIER_ORDER, TIERS

TOOL = "monday-morning-report"
TOOL_VERSION = "1.0"

INK = "0D0D0D"
CHICAGO = "1F2E7A"
RED = "CC100A"
CANVAS = "F5F3EE"


def _norm(s: str) -> str:
    return " ".join(str(s).split()).strip().casefold()


def _load_kpis(read) -> dict[str, str]:
    """Map metric label -> value from a client KPI table (metric,value columns)."""
    cols = {c.lower(): c for c in read.columns}
    metric_c = cols.get("metric") or cols.get("metric_label") or cols.get("label") or read.columns[0]
    value_c = cols.get("value") or (read.columns[1] if len(read.columns) > 1 else None)
    out = {}
    for _, row in read.frame.iterrows():
        m = _norm(row[metric_c])
        if m and value_c is not None:
            out[m] = str(row[value_c]).strip()
    return out


def run(config_path: str, input_path: str | None, out_dir: str, *, final: bool = False) -> dict:
    config = load_config(config_path)
    tier = (config.basis.get("tier") or "").strip()
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    findings: list[Finding] = []
    disclosures: list[str] = []
    input_refs: list[InputRef] = []
    kpis: dict[str, str] = {}

    if tier not in TIERS:
        findings.append(Finding(
            severity="error", category="invalid-tier",
            message=(f"engagement.yml basis.tier {tier!r} is not one of the known tiers: "
                     f"{', '.join(TIER_ORDER)}"),
            spec_ref="INPUT-SPEC §1"))

    if input_path:
        from lailara_engagement import read_table
        read = read_table(input_path)
        input_refs.append(InputRef.from_read_result(read))
        kpis = _load_kpis(read)
        if tier in TIERS:
            expected = [_norm(m["label"]) for m in TIERS[tier]["metrics"]]
            missing = [m["label"] for m in TIERS[tier]["metrics"] if _norm(m["label"]) not in kpis]
            if missing:
                findings.append(Finding(
                    severity="warning", category="missing-metric-values",
                    message=(f"KPI file has no value for {len(missing)} of the tier's metrics "
                             f"(left blank for the client to fill)"),
                    examples=tuple(missing), spec_ref="INPUT-SPEC §2"))
            extra = [k for k in kpis if k not in expected]
            if extra:
                disclosures.append(f"KPI file had {len(extra)} value(s) not in this tier's metric set (ignored)")

    has_error = any(f.severity == "error" for f in findings)
    has_warning = any(f.severity == "warning" for f in findings)
    status = "failed" if has_error else ("warnings" if has_warning else "clean")
    report = PreflightReport(tool=TOOL, status=status, passed=not has_error, findings=findings,
                             disclosures=disclosures, column_mapping={}, n_rows=len(kpis), n_cols=0,
                             spec_version=TOOL_VERSION)
    validation = ("clean" if status == "clean" else
                  f"proceeded with warnings ({report.n_warnings})" if status == "warnings"
                  else "blocked — data not ready")
    provenance = build_provenance(
        tool=TOOL, tool_version=TOOL_VERSION,
        inputs=input_refs or [InputRef(filename="(template — no client values)", sha256="", n_rows=0, n_cols=0)],
        config=config, validation_status=validation)

    if not report.passed:
        paths = write_report(report, config, str(out), provenance=provenance,
                             draft=not final, basename="data-readiness-report",
                             title="Monday Morning Report — Data Readiness")
        return {"status": "blocked", "readiness_report": paths["html"]}

    xlsx_path = out / f"{config.engagement_id}-monday-morning-report.xlsx"
    _build_client_workbook(config, tier, kpis, provenance, validation, draft=not final, path=xlsx_path)
    return {"status": "ok", "tier": tier, "workbook": str(xlsx_path),
            "populated": sum(1 for m in TIERS[tier]["metrics"] if _norm(m["label"]) in kpis),
            "n_warnings": report.n_warnings}


def _build_client_workbook(config, tier, kpis, provenance, validation, *, draft, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "This Week"

    title = f"Monday Morning Report — {config.client_name}"
    if draft:
        title = "[DRAFT] " + title
    ws["A1"] = title
    ws["A1"].font = Font(name="Playfair Display", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=CHICAGO)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"{tier}  ·  Engagement {config.engagement_id}  ·  As of {config.as_of_date.isoformat()}"
    ws["A2"].font = Font(name="Source Sans 3", size=11, bold=True, color="333333")
    ws.merge_cells("A2:E2")

    headers = ["Metric", "Your value", "What it is", "Where to find it", "Watch for / gotcha"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(name="Source Sans 3", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=CHICAGO)

    row = 5
    for m in TIERS[tier]["metrics"]:
        val = kpis.get(_norm(m["label"]), "")
        ws.cell(row=row, column=1, value=m["label"]).font = Font(name="Source Sans 3", size=11, bold=True, color=INK)
        vc = ws.cell(row=row, column=2, value=val)
        vc.font = Font(name="Source Sans 3", size=11, color=CHICAGO, bold=True)
        vc.fill = PatternFill("solid", fgColor="FFF6D6" if not val else "FFFFFF")
        ws.cell(row=row, column=3, value=f"{m['description']}\n({DEFAULT_NOTE})")
        ws.cell(row=row, column=4, value=f"{m['source_system']} — {m['source_detail']}")
        ws.cell(row=row, column=5, value=f"Watch: {m['watch_for']}\nGotcha: {m['gotcha']}")
        for col in range(3, 6):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=col).font = Font(name="Source Sans 3", size=9, color="595959")
        ws.row_dimensions[row].height = 64
        row += 1

    for col, w in zip("ABCDE", (26, 16, 40, 48, 52)):
        ws.column_dimensions[col].width = w

    # Footer + provenance
    ws.oddFooter.left.text = "Monday Morning Report — lailarallc.com"
    ws.oddFooter.left.size = 9
    ws.oddFooter.left.font = "Source Sans 3"
    ws.oddFooter.right.text = "Page &P of &N"

    pv = wb.create_sheet("Provenance")
    lines = [
        ("Client", config.client_name),
        ("Engagement", config.engagement_id),
        ("Prepared by", config.prepared_by),
        ("As of", config.as_of_date.isoformat()),
        ("Tool", f"{TOOL} v{TOOL_VERSION}"),
        ("Config hash", config.config_hash_short),
        ("Validation", validation),
    ]
    for i, ref in enumerate(provenance.inputs):
        if ref.filename and ref.sha256:
            lines.append((f"Input {i+1}", f"{ref.filename}  sha256={ref.sha256[:16]}…  rows={ref.n_rows}"))
    for r, (k, v) in enumerate(lines, start=1):
        pv.cell(row=r, column=1, value=k).font = Font(name="Source Sans 3", size=10, bold=True, color="595959")
        pv.cell(row=r, column=2, value=v).font = Font(name="Source Sans 3", size=10, color=INK)
    pv.column_dimensions["A"].width = 16
    pv.column_dimensions["B"].width = 70

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(prog="monday-morning-report client mode")
    ap.add_argument("--config", required=True)
    ap.add_argument("--input", default=None, help="optional client KPI file (metric,value)")
    ap.add_argument("--out", default="client-output")
    ap.add_argument("--final", action="store_true")
    args = ap.parse_args(argv)
    result = run(args.config, args.input, args.out, final=args.final)
    if result["status"] == "blocked":
        print(f"BLOCKED — data not ready. See {result['readiness_report']}")
        return 3
    print(f"tier {result['tier']}: {result['populated']}/3 metrics populated"
          + (f"; {result['n_warnings']} warning(s)" if result["n_warnings"] else ""))
    print(f"workbook -> {result['workbook']}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
