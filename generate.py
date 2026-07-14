"""
Monday Morning Report — template generator.

Usage:
    python generate.py

Produces: output/monday-morning-report.xlsx

Tabs:
    Setup           — tier selector; metric labels auto-update
    This Week       — 3 input cells, trend arrows, action notes
    History         — 52 pre-created rows; row 3 formula-linked to This Week
    Where to Find   — per-metric data source guide with gotchas
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

from data.metrics import TIERS, TIER_ORDER

OUTPUT_PATH = Path("output") / "monday-morning-report.xlsx"

# Design tokens — Lailara design system
_medium_side = Side(border_style="medium", color="1F2E7A")
INPUT_BORDER = Border(
    left=_medium_side,
    right=_medium_side,
    top=_medium_side,
    bottom=_medium_side,
)

# Trend arrow → (arrow, color) pairs: teal up, red down, grey flat
TREND_STYLES = [("↑", "158F75"), ("↓", "CC100A"), ("→", "666666")]


def build_config_sheet(wb):
    """Hidden sheet that stores tier/metric label data for formula lookup."""
    ws = wb.create_sheet("_Config")
    ws.sheet_state = "hidden"

    # Row 1: headers
    ws["A1"] = "TierKey"
    ws["B1"] = "Metric1"
    ws["C1"] = "Metric2"
    ws["D1"] = "Metric3"

    # Rows 2–4: one row per tier
    for i, tier_key in enumerate(TIER_ORDER, start=2):
        tier = TIERS[tier_key]
        ws.cell(row=i, column=1).value = tier_key
        for j, metric in enumerate(tier["metrics"], start=2):
            ws.cell(row=i, column=j).value = metric["label"]

    return ws


def build_setup_tab(wb):
    """Setup tab: tier selector dropdown + metric labels preview."""
    ws = wb.create_sheet("Setup")

    # --- Header (single cell, wide column spans the visual width) ---
    ws["A1"] = "Monday Morning Report — Setup"
    ws["A1"].font = Font(name="Source Sans 3", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2E7A")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in ["B", "C", "D"]:
        ws[f"{col}1"].fill = PatternFill("solid", fgColor="1F2E7A")
    ws.row_dimensions[1].height = 36

    # --- Instruction ---
    ws["A3"] = "STEP 1 OF 1  —  Click the dropdown below and choose your revenue tier"
    ws["A3"].font = Font(name="Source Sans 3", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill("solid", fgColor="158F75")  # teal instruction bar
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in ["B", "C", "D"]:
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="158F75")
    ws.row_dimensions[3].height = 28

    # --- Tier dropdown ---
    ws["A4"] = "Your tier:"
    ws["A4"].font = Font(name="Source Sans 3", size=11, bold=True, color="333333")
    ws["A4"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[4].height = 32

    # Dropdown cell B4 — navy border + yellow fill signals "editable"
    ws["B4"] = TIER_ORDER[0]  # default
    ws["B4"].font = Font(name="Source Sans 3", size=12, color="1F2E7A", bold=True)
    ws["B4"].fill = PatternFill("solid", fgColor="FEF5D8")
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["B4"].border = INPUT_BORDER

    # "▼ click to change" hint in C4
    ws["C4"] = "▼  click to change"
    ws["C4"].font = Font(name="Source Sans 3", size=10, italic=True, color="1F2E7A")
    ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

    tier_list = ",".join(TIER_ORDER)
    dv = DataValidation(
        type="list",
        formula1=f'"{tier_list}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid tier",
        error="Please select a tier from the list.",
    )
    ws.add_data_validation(dv)
    dv.add(ws["B4"])

    # --- Metric labels preview (formula-driven from _Config + B4) ---
    ws.row_dimensions[5].height = 8  # small gap

    ws["A6"] = "Your three Monday numbers will be:"
    ws["A6"].font = Font(name="Source Sans 3", size=11, bold=True, color="333333")

    for i, col in enumerate(["B", "C", "D"], start=2):
        row = 7
        ws[f"A{row + i - 2}"] = f"Metric {i - 1}:"
        ws[f"A{row + i - 2}"].font = Font(name="Source Sans 3", size=11, color="595959")
        formula = (
            f'=IFERROR(INDEX(_Config!$B$2:$D$4,'
            f'MATCH($B$4,_Config!$A$2:$A$4,0),'
            f'{i - 1}),"—")'
        )
        ws[f"B{row + i - 2}"] = formula
        ws[f"B{row + i - 2}"].font = Font(name="Source Sans 3", size=11, color="1F2E7A", bold=True)

    # --- Default note ---
    ws["A11"] = (
        "These are suggested defaults based on common specialty food brand priorities. "
        "Change any metric label on the 'This Week' tab to track what matters most to your business."
    )
    ws["A11"].font = Font(name="Source Sans 3", size=10, italic=True, color="595959")
    ws["A11"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
    ws.row_dimensions[11].height = 42

    # --- Column widths ---
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    return ws


def build_this_week_tab(wb):
    """This Week tab: 3 metric blocks with input cells, trend arrows, action notes."""
    ws = wb.create_sheet("This Week")

    # --- Header (single cell, remaining cells get matching fill) ---
    ws["A1"] = "Monday Morning Report"
    ws["A1"].font = Font(name="Source Sans 3", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2E7A")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}1"].fill = PatternFill("solid", fgColor="1F2E7A")
    ws.row_dimensions[1].height = 36

    # Week of label + date input
    ws["A2"] = "Week of  →"
    ws["A2"].font = Font(name="Source Sans 3", size=11, bold=True, color="333333")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B2"].number_format = "MMMM D, YYYY"
    ws["B2"].font = Font(name="Source Sans 3", size=11, bold=True, color="1F2E7A")
    ws["B2"].fill = PatternFill("solid", fgColor="FEF5D8")
    ws["B2"].border = INPUT_BORDER
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C2"] = "← type or paste a date"
    ws["C2"].font = Font(name="Source Sans 3", size=9, italic=True, color="B3B3B3")
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26

    # Column headers row
    header_row = 4
    for col, label in [("B", "TYPE HERE  ↓"), ("C", "Last Week (auto)"), ("D", "Trend"), ("E", "Action note")]:
        ws[f"{col}{header_row}"] = label
        bold = col == "B"
        color = "1F2E7A" if col == "B" else "595959"
        ws[f"{col}{header_row}"].font = Font(name="Source Sans 3", size=10, bold=bold, color=color)
        ws[f"{col}{header_row}"].alignment = Alignment(horizontal="center")

    ws[f"A{header_row}"] = "Metric"
    ws[f"A{header_row}"].font = Font(name="Source Sans 3", size=10, bold=True, color="595959")

    # Each metric occupies 2 rows: label+values row, then note row
    # Metric 1: rows 5-6, Metric 2: rows 8-9, Metric 3: rows 11-12
    metric_rows = [5, 8, 11]
    # History col refs for "Last Week" — B/C/D = metric 1/2/3
    history_value_cols = ["B", "C", "D"]

    for idx, start_row in enumerate(metric_rows):
        metric_num = idx + 1
        val_row = start_row
        note_row = start_row + 1

        # Metric label — INDEX/MATCH on _Config driven by Setup tier selection
        formula = (
            f'=IFERROR(INDEX(_Config!$B$2:$D$4,'
            f'MATCH(Setup!$B$4,_Config!$A$2:$A$4,0),'
            f'{metric_num}),"Metric {metric_num}")'
        )
        ws[f"A{val_row}"] = formula
        ws[f"A{val_row}"].font = Font(name="Source Sans 3", size=11, bold=True, color="1F2E7A")
        ws[f"A{val_row}"].alignment = Alignment(vertical="center")

        # This Week input cell — navy border + yellow fill = "type here"
        ws[f"B{val_row}"].fill = PatternFill("solid", fgColor="FEF5D8")
        ws[f"B{val_row}"].font = Font(name="Source Sans 3", size=12, bold=True, color="0D0D0D")
        ws[f"B{val_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{val_row}"].number_format = '#,##0.00'
        ws[f"B{val_row}"].border = INPUT_BORDER

        # Last Week — pulled from History row 3 (the formula-linked mirror of This Week)
        hist_col = history_value_cols[idx]
        ws[f"C{val_row}"] = f"=IFERROR(History!{hist_col}3,\"\")"
        ws[f"C{val_row}"].fill = PatternFill("solid", fgColor="F5F3EE")
        ws[f"C{val_row}"].font = Font(name="Source Sans 3", size=11, color="595959")
        ws[f"C{val_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{val_row}"].number_format = '#,##0.00'

        # Trend arrow formula
        ws[f"D{val_row}"] = (
            f'=IF(ISNUMBER(B{val_row})*ISNUMBER(C{val_row}),'
            f'IF(B{val_row}>C{val_row},"↑",IF(B{val_row}<C{val_row},"↓","→")),"—")'
        )
        ws[f"D{val_row}"].font = Font(name="Source Sans 3", size=14, bold=True)
        ws[f"D{val_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Action note input
        ws[f"E{val_row}"] = "What I'm doing about it…"
        ws[f"E{val_row}"].font = Font(name="Source Sans 3", size=11, italic=True, color="B3B3B3")
        ws[f"E{val_row}"].fill = PatternFill("solid", fgColor="FEF5D8")
        ws[f"E{val_row}"].border = INPUT_BORDER
        ws[f"E{val_row}"].alignment = Alignment(vertical="center", indent=1)

        # Note row (single cell, no merge — column widths carry the visual span)
        ws[f"A{note_row}"] = "← Suggested default — change this label on the Setup tab to track what matters most"
        ws[f"A{note_row}"].font = Font(name="Source Sans 3", size=9, italic=True, color="B3B3B3")
        ws[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[note_row].height = 16
        ws.row_dimensions[val_row].height = 28

    # Conditional formatting for trend arrows — color by arrow symbol
    for start_row in metric_rows:
        cell_ref = f"D{start_row}"
        for arrow, color in TREND_STYLES:
            ws.conditional_formatting.add(
                cell_ref,
                FormulaRule(
                    formula=[f'{cell_ref}="{arrow}"'],
                    font=Font(name="Source Sans 3", size=14, bold=True, color=color),
                ),
            )

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 40

    return ws


def build_history_tab(wb):
    """History tab: 52 pre-created rows, row 3 formula-linked to This Week."""
    ws = wb.create_sheet("History")

    # --- Instruction block (single cell, remaining cells get no content) ---
    ws["A1"] = (
        "Each Monday: after completing 'This Week', copy the values from row 3 "
        "and paste as values into the next blank row below. This builds your trend history over time."
    )
    ws["A1"].font = Font(name="Source Sans 3", size=10, italic=True, color="595959")
    ws["A1"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    # --- Column headers (row 2) ---
    headers = [
        "Date",
        "Metric 1 Value", "Metric 1 Action",
        "Metric 2 Value", "Metric 2 Action",
        "Metric 3 Value", "Metric 3 Action",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(name="Source Sans 3", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2E7A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(border_style="medium", color="0D0D0D"))
    ws.row_dimensions[2].height = 24

    # Row 3: formula-linked to This Week (auto-shows current week's data)
    # This Week: B2=date, B5/B8/B11=metric values, E5/E8/E11=action notes
    linked_formulas = [
        "='This Week'!B2",   # Date
        "='This Week'!B5",   # Metric 1 value
        "='This Week'!E5",   # Metric 1 action
        "='This Week'!B8",   # Metric 2 value
        "='This Week'!E8",   # Metric 2 action
        "='This Week'!B11",  # Metric 3 value
        "='This Week'!E11",  # Metric 3 action
    ]
    for col_idx, formula in enumerate(linked_formulas, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = formula
        cell.font = Font(name="Source Sans 3", size=11, color="595959", italic=True)
        cell.fill = PatternFill("solid", fgColor="F5F3EE")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # Rows 4–54: 51 blank input rows (weeks 2–52)
    # Style objects hoisted out of the loop — only fill alternates per row
    row_font = Font(name="Source Sans 3", size=11, color="333333")
    row_align = Alignment(horizontal="center", vertical="center")
    row_border = Border(bottom=Side(border_style="hair", color="D9D9D9"))
    even_fill = PatternFill("solid", fgColor="FFFFFF")
    odd_fill = PatternFill("solid", fgColor="F5F3EE")

    for row_num in range(4, 55):
        fill = even_fill if row_num % 2 == 0 else odd_fill
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.font = row_font
            cell.alignment = row_align
            cell.border = row_border
            if col_idx == 1:
                cell.number_format = "MMMM D, YYYY"
        ws.row_dimensions[row_num].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 32

    return ws


def build_reference_tab(wb):
    """Where to Find These Numbers — per-metric source guide with gotchas."""
    ws = wb.create_sheet("Where to Find These")

    # --- Header (single cell, col B gets matching fill) ---
    ws["A1"] = "Where to Find These Numbers"
    ws["A1"].font = Font(name="Source Sans 3", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2E7A")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["B1"].fill = PatternFill("solid", fgColor="1F2E7A")
    ws.row_dimensions[1].height = 36

    ws["A2"] = (
        "For each default metric, this tab tells you exactly where to look and what to watch for. "
        "If you've changed a metric label on the Setup tab, refer to that metric's own data source."
    )
    ws["A2"].font = Font(name="Source Sans 3", size=10, italic=True, color="595959")
    ws["A2"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
    ws.row_dimensions[2].height = 36

    current_row = 4

    for tier_key in TIER_ORDER:
        tier = TIERS[tier_key]

        # Tier section header (single cell, col B gets matching fill)
        ws[f"A{current_row}"] = f"Tier: {tier['label']} — {tier['range_note']}"
        ws[f"A{current_row}"].font = Font(name="Source Sans 3", size=12, bold=True, color="FFFFFF")
        ws[f"A{current_row}"].fill = PatternFill("solid", fgColor="1F2E7A")
        ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws[f"B{current_row}"].fill = PatternFill("solid", fgColor="1F2E7A")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for metric in tier["metrics"]:
            # Metric label
            ws[f"A{current_row}"] = metric["label"]
            ws[f"A{current_row}"].font = Font(name="Source Sans 3", size=12, bold=True, color="1F2E7A")
            ws[f"A{current_row}"].alignment = Alignment(vertical="top")
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # Content rows
            content_rows = [
                ("Where to find it:", f"{metric['source_system']} → {metric['source_detail']}"),
                ("What to watch for:", metric["watch_for"]),
                ("Common mistake:", metric["gotcha"]),
            ]

            for label, content in content_rows:
                ws[f"A{current_row}"] = label
                ws[f"A{current_row}"].font = Font(name="Source Sans 3", size=10, bold=True, color="595959")
                ws[f"A{current_row}"].alignment = Alignment(vertical="top")

                ws[f"B{current_row}"] = content
                ws[f"B{current_row}"].font = Font(name="Source Sans 3", size=10, color="333333")
                ws[f"B{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[current_row].height = 42
                current_row += 1

            # Spacer row between metrics
            ws.row_dimensions[current_row].height = 8
            current_row += 1

        # Spacer between tiers
        ws.row_dimensions[current_row].height = 12
        current_row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72

    return ws


def apply_global_styles(wb):
    """Apply Lailara design system styles and print setup across all visible sheets."""
    visible_sheets = ["Setup", "This Week", "History", "Where to Find These"]

    for sheet_name in visible_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Hide gridlines
        ws.sheet_view.showGridLines = False

        # Print setup: letter, 0.6in margins, footer
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.page_margins = PageMargins(
            left=0.6, right=0.6, top=0.75, bottom=0.75,
            header=0.3, footer=0.3
        )

        ws.oddFooter.left.text = "Monday Morning Report — lailara.com"
        ws.oddFooter.left.size = 9
        ws.oddFooter.left.font = "Source Sans 3"
        ws.oddFooter.right.text = "Page &P of &N"
        ws.oddFooter.right.size = 9
        ws.oddFooter.right.font = "Source Sans 3"


def main():
    wb = Workbook()

    # Remove default sheet
    del wb["Sheet"]

    # Build sheets in display order
    build_config_sheet(wb)
    build_setup_tab(wb)
    build_this_week_tab(wb)
    build_history_tab(wb)
    build_reference_tab(wb)

    # Global styling pass
    apply_global_styles(wb)

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Template written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
