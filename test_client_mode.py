"""Client-mode tests for monday-morning-report.

Adversarial fixtures per checklist §6: a template run (no KPI file), a populated
run with a missing metric (proceeds with a warning), an invalid tier (blocked),
the --final watermark drop, and the provenance sheet. Fictional-placeholder
client identity only.

Skipped if lailara_engagement isn't installed.
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

pytest.importorskip("lailara_engagement")

import client_mode  # noqa: E402
from data.metrics import TIERS, TIER_ORDER  # noqa: E402

TIER = TIER_ORDER[1]  # "$10M–$15M"
LABELS = [m["label"] for m in TIERS[TIER]["metrics"]]

_CONFIG = f"""
client: {{name: Meridian Farms}}
engagement: {{id: MER-2026-08}}
as_of_date: "2026-01-31"
demo: true
basis: {{tier: "{TIER}"}}
"""

_BAD_TIER = """
client: {name: Meridian Farms}
engagement: {id: MER-2026-08}
as_of_date: "2026-01-31"
demo: true
basis: {tier: "$50M–$100M"}
"""


def _cfg(tmp_path, text):
    p = tmp_path / "engagement.yml"
    p.write_text(text, encoding="utf-8")
    return str(p)


def _kpi_file(tmp_path, rows):
    p = tmp_path / "kpis.csv"
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "value"])
        w.writerows(rows)
    return str(p)


def test_template_run_no_kpi(tmp_path):
    cfg = _cfg(tmp_path, _CONFIG)
    result = client_mode.run(cfg, None, str(tmp_path / "out"))
    assert result["status"] == "ok"
    assert result["populated"] == 0
    wb = openpyxl.load_workbook(result["workbook"])
    assert wb.sheetnames == ["This Week", "Provenance"]
    ws = wb["This Week"]
    assert ws["A1"].font.name == "Playfair Display"
    assert "Meridian Farms" in ws["A1"].value
    assert ws["A1"].value.startswith("[DRAFT]")
    assert "lailarallc.com" in ws.oddFooter.left.text


def test_populated_run_missing_metric_warns(tmp_path):
    cfg = _cfg(tmp_path, _CONFIG)
    kpis = _kpi_file(tmp_path, [[LABELS[0], "1,250,000"], [LABELS[1], "430,000"]])
    result = client_mode.run(cfg, kpis, str(tmp_path / "out"))
    assert result["status"] == "ok"
    assert result["populated"] == 2
    assert result["n_warnings"] >= 1                     # the 3rd metric is missing
    ws = openpyxl.load_workbook(result["workbook"])["This Week"]
    assert ws["B5"].value == "1,250,000"
    assert ws["B6"].value == "430,000"


def test_header_and_metrics_track_tier_and_as_of_not_hardcoded(tmp_path):
    """This is a fill-in template — the money numbers are user-entered, so there
    is no tool-computed money-over-window to label. Its one data-dependent label
    surface is the 'This Week' header ('{tier} · Engagement {id} · As of {date}')
    plus the tier-appropriate metric labels. Both must track config, not a
    hardcoded tier/date.

    Both halves: feed a distinctive tier + as_of and assert the header + that
    tier's metric labels render, AND assert the demo default tier + date are
    absent."""
    tier_a = TIER_ORDER[0]
    labels_a = [m["label"] for m in TIERS[tier_a]["metrics"]]
    text = ('client: {name: Meridian Farms}\nengagement: {id: MER-2026-08}\n'
            'as_of_date: "2099-09-09"\ndemo: true\n' + f'basis: {{tier: "{tier_a}"}}\n')
    result = client_mode.run(_cfg(tmp_path, text), None, str(tmp_path / "out"))
    assert result["status"] == "ok"
    ws = openpyxl.load_workbook(result["workbook"])["This Week"]
    cells = " ".join(str(ws.cell(row=r, column=c).value)
                     for r in range(1, 25) for c in range(1, 4))
    assert tier_a in cells and "2099-09-09" in cells       # header tracks config
    assert labels_a[0] in cells                            # tier-appropriate metric rendered
    assert TIER not in cells                               # demo default tier must not survive
    assert "2026-01-31" not in cells


def test_invalid_tier_blocks(tmp_path):
    cfg = _cfg(tmp_path, _BAD_TIER)
    result = client_mode.run(cfg, None, str(tmp_path / "out"))
    assert result["status"] == "blocked"
    rr = open(result["readiness_report"], encoding="utf-8").read()
    assert "tier" in rr.lower()
    assert TIER_ORDER[0] in rr                            # names the valid tiers


def test_final_drops_draft(tmp_path):
    cfg = _cfg(tmp_path, _CONFIG)
    result = client_mode.run(cfg, None, str(tmp_path / "out"), final=True)
    ws = openpyxl.load_workbook(result["workbook"])["This Week"]
    assert not ws["A1"].value.startswith("[DRAFT]")


def test_provenance_sheet(tmp_path):
    cfg = _cfg(tmp_path, _CONFIG)
    result = client_mode.run(cfg, None, str(tmp_path / "out"))
    pv = openpyxl.load_workbook(result["workbook"])["Provenance"]
    text = " ".join(str(pv.cell(row=r, column=c).value)
                    for r in range(1, 10) for c in (1, 2))
    assert "Meridian Farms" in text
    assert "Config hash" in text
    assert "monday-morning-report" in text
